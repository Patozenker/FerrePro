"""
PriceScraper — módulo de extracción de precios
Soporta: sitios web (HTML), archivos PDF, archivos Excel/CSV
Exporta a: Excel (.xlsx), CSV
"""

import re
import csv
import time
import json
import logging
import asyncio
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse, urljoin

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import pdfplumber

# Playwright para scroll infinito
PLAYWRIGHT_AVAILABLE = False
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    pass  # Playwright no disponible

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("PriceScraper")

if not PLAYWRIGHT_AVAILABLE:
    log.warning("Playwright no instalado - scroll infinito no disponible")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-AR,es;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

PRICE_RE = re.compile(
    r"""
    (?:AR?\$|USD?\$?|\$|€|£|MXN|CLP|BRL|PYG|UYU|BOB|PEN|COP|VEF)   # símbolo de moneda
    \s*
    [\d.,]+                                                              # número
    |
    [\d.,]+                                                              # número
    \s*
    (?:AR?\$|USD?\$?|\$|€|£)                                            # símbolo después
    """,
    re.VERBOSE,
)

# ─────────────────────────────────────────────────────────────
#  SELECTORES POR SITIO  (extendé según necesites)
# ─────────────────────────────────────────────────────────────
SITE_SELECTORS = {
    "mercadolibre.com.ar": {
        "items":    "li.ui-search-layout__item",
        "title":    ".ui-search-item__title",
        "price":    ".andes-money-amount__fraction",
        "cents":    ".andes-money-amount__cents",
        "currency": ".andes-money-amount__currency-symbol",
        "link":     "a.ui-search-item__group__element",
        "image":    "img.ui-search-result-image__element",
    },
    "mercadolibre.com": {
        "items":    "li.ui-search-layout__item",
        "title":    ".ui-search-item__title",
        "price":    ".andes-money-amount__fraction",
        "cents":    ".andes-money-amount__cents",
        "currency": ".andes-money-amount__currency-symbol",
        "link":     "a.ui-search-item__group__element",
        "image":    "img.ui-search-result-image__element",
    },
    "amazon.com": {
        "items":    "div[data-component-type='s-search-result']",
        "title":    "h2.a-size-base",
        "price":    "span.a-price-whole",
        "cents":    "span.a-price-fraction",
        "currency": "span.a-price-symbol",
        "link":     "a.a-link-normal",
        "image":    "img.s-image",
    },
    "amazon.com.ar": {
        "items":    "div[data-component-type='s-search-result']",
        "title":    "h2.a-size-base",
        "price":    "span.a-price-whole",
        "cents":    "span.a-price-fraction",
        "currency": "span.a-price-symbol",
        "link":     "a.a-link-normal",
        "image":    "img.s-image",
    },
    "falabella.com": {
        "items":    "div.product-grid-item, li.product-item",
        "title":    "div.product-info a, h4.product-name",
        "price":    "span.price-mark, div.price",
        "link":     "a.product-link",
        "image":    "img.product-image",
    },
    "linio.com": {
        "items":    "div.product-item, li.product-item",
        "title":    "div.product-info a, h4.product-name",
        "price":    "span.price-mark, div.price",
        "link":     "a.product-link",
        "image":    "img.product-image",
    },
    "compragamer.com": {
        "items":    "tr[itemprop=itemListElement], div.product-row",
        "title":    "td.product-name, div.product-name",
        "price":    "span.price, td.price",
        "link":     "a[href*=producto]",
        "image":    "img",
    },
    "argseguridad.com": {
        "items":    "div.product-price-and-shipping",
        "title":    "h5, h4, a[href*=product], .product-name",
        "price":    ".product-price, .price, [class*=price]",
        "currency": ".currency, .product-price",
    },
    "pintureriasimagen.com.ar": {
        "items":    "li.product-item",
        "title":    "a.product-item-link, .product-item-name a",
        "price":    ".price, .price-wrapper",
        "link":     "a.product-item-link",
        "image":    "img.product-image-photo",
    },
}


def _parse_price(raw: str) -> float:
    """Convierte formatos: 1.234,56 / 1,234.56 / 1.500 / 99,90 → float"""
    raw = raw.strip().replace("\xa0", "").replace("\u202f", "").replace(" ", "")

    # Eliminar prefijos de moneda (ARS, USD, $, etc.)
    raw = re.sub(r'^(AR?\$|USD?\$?|\$|\u20ac|\u00a3|MXN|CLP|BRL|PYG|UYU|BOB|PEN|COP|VEF)\s*', '', raw, flags=re.I)
    if "," in raw and "." in raw:
        # El último separador es el decimal
        if raw.rfind(",") > raw.rfind("."):  # europeo: 1.234,56
            raw = raw.replace(".", "").replace(",", ".")
        else:                                  # americano: 1,234.56
            raw = raw.replace(",", "")
    elif "," in raw:
        parts = raw.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            raw = raw.replace(",", ".")  # decimal: 99,90
        else:
            raw = raw.replace(",", "")   # miles: 1,000
    elif "." in raw:
        parts = raw.split(".")
        if len(parts) == 2 and len(parts[1]) == 3:
            raw = raw.replace(".", "")   # miles argentino: 1.500 → 1500
        # else: decimal normal 1.50
    try:
        return float(re.sub(r"[^\d.]", "", raw))
    except ValueError:
        return 0.0


def _get(url: str, timeout: int = 20) -> Optional[requests.Response]:
    """GET con reintentos y backoff."""
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            log.warning(f"Intento {attempt+1}/3 fallido para {url}: {e}")
            if attempt < 2:
                time.sleep(2 ** attempt + 1)
    log.error(f"Error definitivo para {url}")
    return None


# ─────────────────────────────────────────────────────────────
#  SCRAPING CON PLAYWRIGHT (SCROLL INFINITO)
# ─────────────────────────────────────────────────────────────

def _scrape_with_scroll(url: str, scroll_pause: float = 2.0, max_scrolls: int = 50) -> str:
    """
    Usa Playwright para cargar página y hacer scroll infinito.
    Retorna el HTML completo después de cargar todo.
    max_scrolls: cantidad máxima de scrolls (default 50 para ~35+ páginas)
    """
    if not PLAYWRIGHT_AVAILABLE:
        log.warning("Playwright no disponible")
        return ""

    html = ""
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_default_timeout(30000)

            log.info(f"Cargando {url} con Playwright...")
            page.goto(url, wait_until="networkidle")

            # Scroll progresivo para activar carga infinita
            last_height = 0
            for scroll_num in range(max_scrolls):
                # Scroll hasta el fondo
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(scroll_pause)

                # Verificar si hay botón "Cargar más" o similar
                try:
                    load_more = page.query_selector("button:has-text('Cargar más'), button:has-text('Ver más'), button:has-text('Más resultados'), [class*=load-more], [class*=infinite]")
                    if load_more:
                        load_more.click()
                        time.sleep(scroll_pause)
                except:
                    pass

                # Verificar si új content was loaded
                new_height = page.evaluate("document.body.scrollHeight")
                if new_height == last_height:
                    log.info(f"No hay más contenido después de {scroll_num + 1} scrolls")
                    break
                last_height = new_height
                log.info(f"Scroll {scroll_num + 1}/{max_scrolls} - altura: {new_height}")

            html = page.content()
            browser.close()
            log.info("Playwright: HTML extraído")

    except Exception as e:
        log.error(f"Error con Playwright: {e}")

    return html


# ─────────────────────────────────────────────────────────────
#  SCRAPING WEB
# ─────────────────────────────────────────────────────────────

def scrape_generic_with_html(url: str, html: str) -> list[dict]:
    """
    Scraper genérico que recibe HTML en lugar de URL.
    Usado cuando Playwright genera el HTML con scroll infinito.
    """
    soup = BeautifulSoup(html, "lxml")
    results = []

    # Resto del código de scrape_generic...
    from urllib.parse import urlparse
    domain = urlparse(url).netloc
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    # (same logic as scrape_generic from here)
    seen = set()
    CARD_CLASSES = re.compile(r"(product|item|card|listing|resultado|kit|articul|oferta|prod|grid|row)", re.I)
    candidates = []
    for tag in soup.find_all(["li", "article", "div", "tr"]):
        cls = " ".join(tag.get("class", []))
        if CARD_CLASSES.search(cls):
            text = tag.get_text(" ", strip=True)
            if PRICE_RE.search(text) and len(text) > 10:
                candidates.append(tag)

    if len(candidates) < 3:
        for tag in soup.find_all(["li", "article", "tr"]):
            text = tag.get_text(" ", strip=True)
            if PRICE_RE.search(text) and 10 < len(text) < 2000:
                candidates.append(tag)

    if len(candidates) < 3:
        for tag in soup.find_all(["div", "li"]):
            text = tag.get_text(" ", strip=True)
            if PRICE_RE.search(text):
                words = text.split()
                if len(words) >= 2:
                    candidates.append(tag)

    if len(candidates) < 3:
        price_els = soup.find_all(class_=re.compile(r'price', re.I))
        for price_el in price_els:
            parent = price_el.parent
            depth = 0
            while parent and depth < 10:
                cls = " ".join(parent.get("class", []))
                text = parent.get_text(" ", strip=True)
                if len(text) > 30 and PRICE_RE.search(text):
                    candidates.append(parent)
                    break
                parent = parent.parent
                depth += 1

    log.info(f"Candidatos a tarjetas de producto: {len(candidates)}")

    for card in candidates:
        text = card.get_text(" ", strip=True)

        # Extraer nombre
        title = ""
        for tag in ["h1","h2","h3","h4","h5","strong","b"]:
            el = card.find(tag)
            if el:
                t = el.get_text(strip=True)
                if t and len(t) > 3:
                    title = t[:200]
                    break
        if not title:
            a = card.find("a", title=True)
            if a:
                title = a.get("title","").strip()[:200]
        if not title:
            a = card.find("a")
            if a and a.get_text(strip=True):
                title = a.get_text(strip=True)[:200]
        if not title:
            title = text[:100]

        # Limpiar título
        title = re.sub(r'\d{1,3}\.\d{3},\d{2}\s*(?:USD|ARS)?', '', title)
        title = re.sub(r'USD\s*[\d.,]+', '', title)
        title = re.sub(r'ARS\s*[\d.,]+', '', title)
        title = re.sub(r'\$[\d.,]+', '', title)
        title = title.strip()

        # Filtrar no-productos
        non_product_keywords = ['cuenta', 'cliente', 'login', 'registro', 'carrito', 'comprar', 'iniciar sesión', 'mi cuenta']
        title_lower = title.lower()
        if any(kw in title_lower for kw in non_product_keywords):
            continue

        # Extraer link
        link_el = card.find("a", href=True)
        link = urljoin(url, link_el["href"]) if link_el else url

        # Extraer imagen
        img_el = card.find("img")
        image = ""
        if img_el:
            image = img_el.get("src") or img_el.get("data-src") or img_el.get("data-load") or ""
            if image and not image.startswith(("http://", "https://", "data:")):
                image = urljoin(url, image)

        # Extraer precios
        prices = []
        for pm in PRICE_RE.finditer(text):
            raw = pm.group()
            val = _parse_price(re.sub(r"[^\d.,]", "", raw))
            if val > 0:
                prices.append((val, raw.strip()))

        # Estrategia 2: ARS patterns
        ars_pattern = re.compile(r'ARS\s*([\d.]+,?\d*)', re.I)
        for m in ars_pattern.finditer(text):
            raw_num = m.group(1)
            val = _parse_price(raw_num)
            if val > 0:
                prices.append((val, f"ARS {raw_num}"))

        # Estrategia 3: large argentine numbers
        large_arg = re.compile(r'([\d]{1,3}\.[\d]{3},\d{2})')
        for m in large_arg.finditer(text):
            raw_num = m.group(1)
            val = _parse_price(raw_num)
            if val > 0:
                prices.append((val, raw_num))

        if not prices:
            continue

        # Priorizar ARS
        price_val, price_raw = prices[0]
        precio_usd = precio_ars = 0.0
        raw_usd = raw_ars = ""
        for val, raw in prices:
            if re.search(r"ARS|pesos", raw, re.I):
                precio_ars = val; raw_ars = raw
            elif re.search(r"USD|U\$|dólar", raw, re.I):
                if not precio_usd or val < precio_usd:
                    precio_usd = val; raw_usd = raw

        if precio_ars > 0:
            price_val = precio_ars
            price_raw = raw_ars or prices[0][1]

        key = title[:60].lower().strip()
        if key in seen:
            continue
        seen.add(key)

        item = {
            "nombre": title,
            "precio": price_val,
            "precio_raw": price_raw,
            "link": link,
            "imagen": image,
            "fuente": domain,
            "fecha": fecha,
        }
        results.append(item)

    log.info(f"Generic scraper (Playwright) → {len(results)} items de {url}")
    return results


def scrape_generic(url: str) -> list[dict]:
    """
    Scraper genérico mejorado: detecta tarjetas de producto completas
    buscando contenedores que tengan tanto título como precio.
    """
    r = _get(url)
    if not r:
        return []

    soup = BeautifulSoup(r.text, "lxml")
    results = []
    seen = set()
    domain = urlparse(url).netloc
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Estrategia 1: buscar tarjetas de producto (elementos repetidos con precio) ──
    # Candidatos: li, article, div con clases típicas de producto
    CARD_CLASSES = re.compile(
        r"(product|item|card|listing|resultado|kit|articul|oferta|prod|grid|row)",
        re.I
    )
    candidates = []
    for tag in soup.find_all(["li", "article", "div", "tr"]):
        cls = " ".join(tag.get("class", []))
        if CARD_CLASSES.search(cls):
            # Solo tarjetas con precio adentro
            text = tag.get_text(" ", strip=True)
            if PRICE_RE.search(text) and len(text) > 10:
                candidates.append(tag)

    # Si no hay candidatos por clase, buscar por estructura repetida
    if len(candidates) < 3:
        for tag in soup.find_all(["li", "article", "tr"]):
            text = tag.get_text(" ", strip=True)
            if PRICE_RE.search(text) and 10 < len(text) < 2000:
                candidates.append(tag)

    # Estrategia 1.5: buscar cualquier elemento con precio que parezca producto
    if len(candidates) < 3:
        for tag in soup.find_all(["div", "li"]):
            text = tag.get_text(" ", strip=True)
            # Buscar patrones de precio + posible nombre
            if PRICE_RE.search(text):
                # Verificar que tenga al menos 2 palabras y un precio razonable
                words = text.split()
                if len(words) >= 2:
                    candidates.append(tag)

    # Estrategia 1.6: buscar por elementos con precio específico (para sitios como argseguridad)
    if len(candidates) < 3:
        # Buscar elementos con clase que contenga 'price'
        price_els = soup.find_all(class_=re.compile(r'price', re.I))
        for price_el in price_els:
            # Buscar el contenedor del producto subiendo en el árbol
            parent = price_el.parent
            depth = 0
            while parent and depth < 10:
                cls = " ".join(parent.get("class", []))
                text = parent.get_text(" ", strip=True)
                # Verificar que el contenedor tenga suficiente contenido (no solo precio)
                if len(text) > 30 and PRICE_RE.search(text):
                    candidates.append(parent)
                    break
                parent = parent.parent
                depth += 1

    log.info(f"Candidatos a tarjetas de producto: {len(candidates)}")

    for card in candidates:
        text = card.get_text(" ", strip=True)

        # ── Extraer nombre (primero del titulo/heading/link) ──
        title = ""
        for tag in ["h1","h2","h3","h4","h5","strong","b"]:
            el = card.find(tag)
            if el:
                t = el.get_text(strip=True)
                if t and len(t) > 3:
                    title = t[:200]
                    break
        if not title:
            a = card.find("a", title=True)
            if a:
                title = a.get("title","").strip()[:200]
        if not title:
            a = card.find("a")
            if a and a.get_text(strip=True):
                title = a.get_text(strip=True)[:200]
        if not title:
            title = text[:100]

        # Limpiar título: quitar patrones de precios mezclados
        title = re.sub(r'\d{1,3}\.\d{3},\d{2}\s*(?:USD|ARS)?', '', title)
        title = re.sub(r'USD\s*[\d.,]+', '', title)
        title = re.sub(r'ARS\s*[\d.,]+', '', title)
        title = re.sub(r'\$[\d.,]+', '', title)
        title = title.strip()

        # Filtrar títulos que no parecen productos reales
        non_product_keywords = ['cuenta', 'cliente', 'login', 'registro', 'carrito', 'comprar', 'iniciar sesión', 'mi cuenta']
        title_lower = title.lower()
        if any(kw in title_lower for kw in non_product_keywords):
            continue

        # ── Extraer código / SKU ──
        code = ""
        code_re = re.compile(r"\b([A-Z][A-Z0-9\-]{3,20})\b")
        m = code_re.search(text)
        if m:
            code = m.group(1)

        # ── Extraer link ──
        link_el = card.find("a", href=True)
        link = urljoin(url, link_el["href"]) if link_el else url

        # ── Extraer imagen ──
        img_el = card.find("img")
        image = ""
        if img_el:
            image = img_el.get("src") or img_el.get("data-src") or img_el.get("data-load") or ""
            # Si es URL relativa, completar con el dominio
            if image and not image.startswith(("http://", "https://", "data:")):
                image = urljoin(url, image)

        # ── Extraer TODOS los precios del card ──
        prices = []

        # Estrategia 1: buscar con regex estándar
        for pm in PRICE_RE.finditer(text):
            raw = pm.group()
            val = _parse_price(re.sub(r"[^\d.,]", "", raw))
            if val > 0:
                prices.append((val, raw.strip()))

        # Estrategia 2: buscar patrones "ARS 94.745,19" o "94.745,19 ARS"
        ars_pattern = re.compile(r'ARS\s*([\d.]+,?\d*)', re.I)
        for m in ars_pattern.finditer(text):
            raw_num = m.group(1)
            val = _parse_price(raw_num)
            if val > 0:
                prices.append((val, f"ARS {raw_num}"))

        # Estrategia 3: buscar números grandes argentino (formato con punto de miles)
        # como "94.745,19" - detectar por la presencia de coma al final
        large_arg = re.compile(r'([\d]{1,3}\.[\d]{3},\d{2})')
        for m in large_arg.finditer(text):
            raw_num = m.group(1)
            val = _parse_price(raw_num)
            if val > 0:
                prices.append((val, raw_num))

        if not prices:
            continue

        # Precio principal = Priorizar ARS (pesos argentinos) sobre USD
        price_val, price_raw = prices[0]

        # Detectar moneda: buscar primero ARS que es lo que necesitamos
        precio_usd = precio_ars = 0.0
        raw_usd = raw_ars = ""
        for val, raw in prices:
            # Primero buscar ARS (prioridad para precios locales)
            if re.search(r"ARS|pesos", raw, re.I):
                precio_ars = val; raw_ars = raw
            # Después USD
            elif re.search(r"USD|U\$|dólar", raw, re.I):
                if not precio_usd or val < precio_usd:  # tomar el menor USD
                    precio_usd = val; raw_usd = raw

        # Si hay precio ARS, usarlo como principal; sino usar el primero
        if precio_ars > 0:
            price_val = precio_ars
            price_raw = raw_ars or prices[0][1]

        # Deduplicar por nombre (usar el precio más alto si hay duplicados)
        key = title[:60].lower().strip()
        if key in seen:
            # Ya existe, skipear o actualizar con precio más alto
            continue
        seen.add(key)

        item = {
            "nombre":     title,
            "precio":     precio_ars or price_val,
            "precio_raw": raw_ars or price_raw,
            "link":       link,
            "imagen":     image,
            "fuente":     domain,
            "fecha":      fecha,
        }
        if code:        item["codigo"]     = code
        if precio_usd:  item["precio_usd"] = precio_usd
        if raw_usd:     item["raw_usd"]    = raw_usd
        results.append(item)

    # ── Estrategia 2: tabla HTML (lista de precios en <table>) ──
    if len(results) < 3:
        for table in soup.find_all("table"):
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue
            header = [td.get_text(strip=True).lower() for td in rows[0].find_all(["th","td"])]
            col_n = _find_col(header, ["nombre","producto","descripcion","articulo","item","detalle","description"])
            col_p = _find_col(header, ["precio","importe","valor","price","costo","monto"])
            col_c = _find_col(header, ["codigo","cod","ref","sku","id"])
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all(["td","th"])]
                if not cells: continue
                nombre = cells[col_n] if col_n is not None and col_n < len(cells) else cells[0] if cells else ""
                pr_raw = cells[col_p] if col_p is not None and col_p < len(cells) else ""
                code   = cells[col_c] if col_c is not None and col_c < len(cells) else ""
                if not pr_raw:
                    for c in cells:
                        if PRICE_RE.search(c): pr_raw = c; break
                pr_val = _parse_price(re.sub(r"[^\d.,]","",pr_raw))
                if pr_val <= 0 and not nombre: continue
                key = (nombre[:60], pr_val)
                if key in seen: continue
                seen.add(key)
                results.append({"codigo":code,"nombre":nombre,"precio":pr_val,"precio_raw":pr_raw,"fuente":domain,"fecha":fecha})

    log.info(f"Generic scraper → {len(results)} items de {url}")
    return results


def scrape_site(url: str, use_scroll: bool = False) -> list[dict]:
    """
    Scraper con selectores específicos por sitio.
    Cae a genérico si no hay configuración.
    use_scroll: si True, usa Playwright para scroll infinito.
    """
    domain = urlparse(url).netloc.replace("www.", "")
    sel = None
    for key in SITE_SELECTORS:
        if key in domain:
            sel = SITE_SELECTORS[key]
            break

    # Si se requiere scroll infinito y Playwright está disponible, usar scrape_generic
    if use_scroll and PLAYWRIGHT_AVAILABLE:
        html = _scrape_with_scroll(url)
        if html:
            return scrape_generic_with_html(url, html)
        return []

    if sel is None:
        return scrape_generic(url)

    # Scraper normal sin scroll
    r = _get(url)
    if not r:
        return []
    soup = BeautifulSoup(r.text, "lxml")
    results = []

    for item in soup.select(sel["items"]):
        try:
            title_el  = item.select_one(sel["title"])
            price_el  = item.select_one(sel["price"])
            cents_el  = item.select_one(sel.get("cents", ""))
            curr_el   = item.select_one(sel.get("currency", ""))
            link_el   = item.select_one(sel["link"])
            img_el    = item.select_one(sel.get("image", ""))

            if not title_el or not price_el:
                continue

            title    = title_el.get_text(strip=True)
            price_s  = price_el.get_text(strip=True)
            cents_s  = cents_el.get_text(strip=True) if cents_el else "00"
            currency = curr_el.get_text(strip=True) if curr_el else "$"
            link     = link_el.get("href", url) if link_el else url
            image    = (img_el.get("src") or img_el.get("data-src", "")) if img_el else ""

            price_full = f"{price_s}.{cents_s}" if cents_s not in ("", "00") else price_s
            price_val  = _parse_price(price_full)

            results.append({
                "nombre":    title,
                "precio":    price_val,
                "precio_raw": f"{currency} {price_s},{cents_s}",
                "link":      link,
                "imagen":    image,
                "url":       url,
                "fuente":    domain,
                "fecha":     datetime.now().strftime("%Y-%m-%d %H:%M"),
            })
        except Exception as e:
            log.debug(f"Error parseando item: {e}")

    log.info(f"Site scraper ({domain}) → {len(results)} items")

    # Si no encontró nada con selectors específicos, caer a genérico
    if not results:
        log.info(f"Site scraper no encontró items, cayendo a genérico...")
        return scrape_generic(url)

    return results


# ─────────────────────────────────────────────────────────────
#  PARSEO PDF
# ─────────────────────────────────────────────────────────────

def parse_pdf(source) -> list[dict]:
    """
    Extrae precios de un PDF (path, bytes o URL).
    Estrategia: primero tablas, luego texto libre con regex.
    """
    # Cargar el PDF
    if isinstance(source, (str, Path)):
        src_str = str(source)
        if src_str.startswith("http"):
            r = _get(src_str)
            if not r:
                return []
            pdf_bytes = BytesIO(r.content)
        else:
            pdf_bytes = src_str
    elif isinstance(source, bytes):
        pdf_bytes = BytesIO(source)
    else:
        pdf_bytes = source  # ya es file-like

    results = []
    nombre_origen = getattr(source, "name", "archivo.pdf") if hasattr(source, "name") else str(source)

    with pdfplumber.open(pdf_bytes) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            # 1️⃣  Intentar extraer tablas estructuradas
            for table in page.extract_tables():
                if not table:
                    continue
                # Detectar encabezados
                header = [str(c or "").lower().strip() for c in table[0]]
                col_nombre = _find_col(header, ["nombre","producto","descripcion","articulo","item","detalle","description","name","titulo"])
                col_precio  = _find_col(header, ["precio","importe","valor","price","costo","pvp","p.v.p","monto","neto","lista","list","unit"])
                col_codigo  = _find_col(header, ["codigo","cod","ref","sku","id","art","code","clave"])
                col_stock   = _find_col(header, ["stock","disponible","cantidad","qty","en stock"])
                col_iva     = _find_col(header, ["iva","tax","impuesto","vat"])

                start = 1 if col_nombre is not None or col_precio is not None else 0

                for row in table[start:]:
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue
                    def _s(v): return str(v or "").strip()
                    nombre    = _s(row[col_nombre] if col_nombre is not None and col_nombre < len(row) else "")
                    precio_raw= _s(row[col_precio]  if col_precio  is not None and col_precio  < len(row) else "")
                    codigo    = _s(row[col_codigo]  if col_codigo  is not None and col_codigo  < len(row) else "")
                    stock     = _s(row[col_stock]   if 'col_stock' in dir() and col_stock  is not None and col_stock  < len(row) else "")
                    iva_raw   = _s(row[col_iva]     if 'col_iva'   in dir() and col_iva    is not None and col_iva    < len(row) else "")

                    if not nombre and len(row) > 0:
                        nombre = _s(row[0])[:120]

                    if not precio_raw:
                        for cell in row:
                            m = PRICE_RE.search(str(cell or ""))
                            if m: precio_raw = m.group(); break

                    precio_val = _parse_price(re.sub(r"[^\d.,]", "", precio_raw)) if precio_raw else 0.0
                    precio_iva = 0.0
                    if iva_raw:
                        iva_pct = _parse_price(re.sub(r"[^\d.,]","",iva_raw))
                        if 0 < iva_pct <= 100 and precio_val > 0:
                            precio_iva = round(precio_val*(1+iva_pct/100),2)

                    if precio_val <= 0 and not nombre and not codigo:
                        continue

                    item = {
                        "codigo":    codigo,
                        "nombre":    nombre[:200],
                        "precio":    precio_val,
                        "precio_raw": precio_raw,
                        "pagina":    page_num,
                        "fuente":    Path(nombre_origen).name,
                        "fecha":     datetime.now().strftime("%Y-%m-%d %H:%M"),
                    }
                    if stock:     item["stock"] = stock
                    if iva_raw:   item["iva"]   = iva_raw
                    if precio_iva > 0: item["precio_c_iva"] = precio_iva
                    results.append(item)

            # 2️⃣  Si no hubo tablas, texto libre
            if not results:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    m = PRICE_RE.search(line)
                    if not m:
                        continue
                    precio_raw = m.group()
                    precio_val = _parse_price(re.sub(r"[^\d.,]", "", precio_raw))
                    nombre = line[:line.find(m.group())].strip()[:200] or line[:100]
                    results.append({
                        "codigo":    "",
                        "nombre":    nombre,
                        "precio":    precio_val,
                        "precio_raw": precio_raw,
                        "pagina":    page_num,
                        "fuente":    Path(nombre_origen).name,
                        "fecha":     datetime.now().strftime("%Y-%m-%d %H:%M"),
                    })

    log.info(f"PDF parser → {len(results)} items de {nombre_origen}")
    return results


def _find_col(header: list, candidates: list) -> Optional[int]:
    for i, h in enumerate(header):
        for c in candidates:
            if c in h:
                return i
    return None


# ─────────────────────────────────────────────────────────────
#  PARSEO EXCEL / CSV
# ─────────────────────────────────────────────────────────────

def parse_excel(source) -> list[dict]:
    """
    Extrae precios de un Excel o CSV (path, bytes, URL).
    """
    if isinstance(source, (str, Path)):
        src_str = str(source)
        if src_str.startswith("http"):
            r = _get(src_str)
            if not r:
                return []
            data = BytesIO(r.content)
            name = src_str.split("/")[-1]
        else:
            data = src_str
            name = Path(src_str).name
    elif isinstance(source, bytes):
        data = BytesIO(source)
        name = "archivo.xlsx"
    else:
        data = source
        name = getattr(source, "name", "archivo.xlsx")

    ext = Path(str(name)).suffix.lower()
    try:
        if ext == ".csv":
            # Probar distintos separadores y encodings (común en archivos argentinos/europeos)
            df = None
            # Leer los bytes una vez para poder reintentar con distintos separadores
            if isinstance(data, str):
                _raw_csv = open(data, "rb").read()
            else:
                if hasattr(data, "read"):
                    _raw_csv = data.read()
                else:
                    _raw_csv = data
            log.debug(f"_raw_csv bytes: {len(_raw_csv)}, first 60: {_raw_csv[:60]}")
            for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
                for sep in [";", ",", "\t", "|"]:
                    try:
                        trial = pd.read_csv(BytesIO(_raw_csv), dtype=str,
                                            encoding=enc, sep=sep,
                                            encoding_errors="replace")
                        # Una parse válida tiene más de 1 columna
                        if len(trial.columns) > 1:
                            df = trial
                            log.info(f"CSV leído con sep='{sep}' enc='{enc}' → {len(df)} filas, {len(df.columns)} cols: {list(df.columns[:3])}")
                            break
                    except Exception as _e:
                        log.debug(f"  sep='{sep}' enc='{enc}': {_e}")
                if df is not None:
                    break
            # Fallback: 1 sola columna con cualquier separador
            if df is None:
                for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
                    try:
                        df = pd.read_csv(BytesIO(_raw_csv), dtype=str, encoding=enc,
                                         encoding_errors="replace")
                        log.info(f"CSV fallback enc='{enc}' → {len(df)} filas")
                        break
                    except Exception as _e2:
                        log.debug(f"  fallback enc='{enc}': {_e2}")
            if df is None:
                log.error("No se pudo parsear el CSV con ninguna combinación")
                return []
        else:
            df = pd.read_excel(data, dtype=str, engine="openpyxl" if ext in (".xlsx", ".xlsm") else "xlrd")
    except Exception as e:
        log.error(f"Error leyendo {name}: {e}")
        return []

    # Normalizar columnas
    df.columns = [str(c).lower().strip() for c in df.columns]
    log.info(f"Columnas detectadas: {list(df.columns)}")

    COLS_NOMBRE = ["nombre","producto","descripcion","articulo","item","detalle","description","name","product","titulo","denominacion","articulo","rubro","detallearticulo","art"]
    COLS_PRECIO  = ["precio","importe","valor","price","costo","pvp","p.v.p","monto","unit","neto","lista","list","pvp","precioventa","preciofinal","tarifa","tarifa1"]
    COLS_CODIGO  = ["codigo","cod","ref","sku","id","art","code","clave","part","codarticulo","codprov","cod prov"]
    COLS_STOCK   = ["stock","disponible","cantidad","qty","inventory","existencia","en stock","cant","cantidad","stock actual"]
    COLS_IVA     = ["iva","tax","impuesto","vat","gst","i.v.a"]
    COLS_MARCA   = ["marca","brand","fabricante","manufacturer","proveedor","supplier","fabricante","orig"]
    COLS_CATEG   = ["categoria","category","rubro","familia","group","tipo","linea","familia"]

    col_nombre = _find_col(list(df.columns), COLS_NOMBRE)
    col_precio  = _find_col(list(df.columns), COLS_PRECIO)
    col_codigo  = _find_col(list(df.columns), COLS_CODIGO)
    col_stock   = _find_col(list(df.columns), COLS_STOCK)
    col_iva     = _find_col(list(df.columns), COLS_IVA)
    col_marca   = _find_col(list(df.columns), COLS_MARCA)
    col_categ   = _find_col(list(df.columns), COLS_CATEG)

    # Si no se detectó nombre, usar primera columna de texto
    if col_nombre is None and len(df.columns) > 0:
        col_nombre = 0
        log.info(f"Columna nombre → primera columna: '{df.columns[0]}'")

    # Si no se detectó precio, buscar la primera columna numérica (distinta del nombre)
    if col_precio is None:
        for i, col in enumerate(df.columns):
            if i == col_nombre:
                continue
            sample = df.iloc[:, i].dropna().head(15)
            numeric_count = sum(1 for v in sample if re.search(r'^\s*[\$€£]?[\d.,]+\s*$', str(v).strip()))
            if sample.size > 0 and numeric_count >= len(sample) * 0.5:
                col_precio = i
                log.info(f"Columna precio auto: '{col}' (idx {i})")
                break

    results = []
    def _gv(vals, idx):
        if idx is None or idx >= len(vals): return ""
        v = str(vals[idx] or "").strip()
        return "" if v.lower() in ("nan","none","<na>") else v

    for _, row in df.iterrows():
        vals = list(row.values)
        nombre    = _gv(vals, col_nombre)[:200]
        precio_raw= _gv(vals, col_precio)
        codigo    = _gv(vals, col_codigo)
        stock     = _gv(vals, col_stock)  if col_stock  is not None else ""
        iva_raw   = _gv(vals, col_iva)    if col_iva    is not None else ""
        marca     = _gv(vals, col_marca)  if col_marca  is not None else ""
        categ     = _gv(vals, col_categ)  if col_categ  is not None else ""

        # Skip pure header-looking rows (all values equal their column name)
        if nombre.lower() in [c.lower() for c in df.columns]:
            continue

        if not precio_raw:
            for v in vals:
                m = PRICE_RE.search(str(v or ""))
                if m:
                    precio_raw = m.group()
                    break

        precio_val = _parse_price(re.sub(r"[^\d.,]", "", precio_raw)) if precio_raw else 0.0

        # Calcular precio con IVA si está disponible
        precio_iva = 0.0
        if iva_raw:
            iva_pct = _parse_price(re.sub(r"[^\d.,]", "", iva_raw))
            if 0 < iva_pct <= 100 and precio_val > 0:
                precio_iva = round(precio_val * (1 + iva_pct/100), 2)
            elif iva_pct > 100:  # ya es el monto total
                precio_iva = iva_pct

        if precio_val <= 0 and nombre == "" and codigo == "":
            continue

        item = {
            "codigo":    codigo,
            "nombre":    nombre if nombre else (codigo or f"Item {_+1}"),
            "precio":    precio_val,
            "precio_raw": precio_raw,
        }
        if stock:   item["stock"]  = stock
        if iva_raw: item["iva"]    = iva_raw
        if precio_iva > 0: item["precio_c_iva"] = precio_iva
        if marca:   item["marca"]  = marca
        if categ:   item["categoria"] = categ
        item["fuente"] = name
        item["fecha"]  = datetime.now().strftime("%Y-%m-%d %H:%M")
        results.append(item)

    log.info(f"Excel parser → {len(results)} items de {name}")
    return results


# ─────────────────────────────────────────────────────────────
#  EXPORTACIÓN
# ─────────────────────────────────────────────────────────────

def export_csv(items: list[dict], path: str) -> str:
    if not items:
        return ""
    keys = list(items[0].keys())
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(items)
    log.info(f"CSV exportado → {path}")
    return path


def export_excel(items: list[dict], path: str, sheet_name: str = "Precios") -> str:
    if not items:
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Estilos
    header_fill = PatternFill("solid", fgColor="1E1E2E")
    header_font = Font(name="Calibri", bold=True, color="FF6B35", size=11)
    alt_fill    = PatternFill("solid", fgColor="F5F5FF")
    border = Border(
        bottom=Side(style="thin", color="E0E0E8"),
    )
    center = Alignment(horizontal="center", vertical="center")

    keys = list(items[0].keys())

    # Cabecera
    for col, key in enumerate(keys, 1):
        cell = ws.cell(row=1, column=col, value=key.upper())
        cell.font = header_fill_font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Datos
    for row_num, item in enumerate(items, 2):
        fill = alt_fill if row_num % 2 == 0 else None
        for col, key in enumerate(keys, 1):
            val = item.get(key, "")
            # Convertir precio a número real
            if key == "precio" and isinstance(val, (int, float)):
                val = val
            cell = ws.cell(row=row_num, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.border = border
            if key == "precio":
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # Autowidth
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Hoja de resumen
    ws_sum = wb.create_sheet("Resumen")
    ws_sum["A1"] = "Resumen"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Total productos"
    ws_sum["B3"] = len(items)
    ws_sum["A4"] = "Precio mínimo"
    ws_sum["B4"] = min((i.get("precio", 0) or 0 for i in items), default=0)
    ws_sum["A5"] = "Precio máximo"
    ws_sum["B5"] = max((i.get("precio", 0) or 0 for i in items), default=0)
    ws_sum["A6"] = "Precio promedio"
    prices = [i.get("precio", 0) or 0 for i in items if i.get("precio", 0)]
    ws_sum["B6"] = round(sum(prices) / len(prices), 2) if prices else 0
    ws_sum["A7"] = "Exportado"
    ws_sum["B7"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    for r in ["B3", "B4", "B5", "B6"]:
        ws_sum[r].number_format = '#,##0.00'

    wb.save(path)
    log.info(f"Excel exportado → {path}")
    return path


# ─────────────────────────────────────────────────────────────
#  API UNIFICADA
# ─────────────────────────────────────────────────────────────

def scrape(source: str, source_type: str = "auto", use_scroll: bool = False) -> list[dict]:
    """
    Punto de entrada principal.
    source:      URL, ruta de archivo, o bytes
    source_type: "web" | "pdf" | "excel" | "csv" | "auto"
    use_scroll:  True para usar Playwright (scroll infinito)
    """
    # Resolver file:// URLs como rutas locales
    if isinstance(source, str) and source.startswith("file:///"):
        from urllib.request import url2pathname
        import urllib.parse
        local_path = urllib.parse.unquote(source[8:])  # quitar file:///
        # En Windows: file:///C:/... → C:/...
        # En Linux:   file:///home/... → /home/...
        if not local_path.startswith("/"):
            local_path = "/" + local_path
        source = local_path
        log.info(f"file:// resuelto como ruta local: {source}")

    if source_type == "auto":
        lower = source.lower() if isinstance(source, str) else ""
        if lower.endswith(".pdf") or (lower.endswith(".pdf") and "http" not in lower):
            source_type = "pdf"
        elif any(lower.endswith(x) for x in (".xlsx", ".xls", ".xlsm", ".ods")):
            source_type = "excel"
        elif lower.endswith(".csv"):
            source_type = "csv"
        elif isinstance(source, str) and not source.startswith("http"):
            # Es una ruta local — detectar por extensión
            ext = Path(source).suffix.lower() if Path(source).exists() else ""
            if ext == ".pdf":     source_type = "pdf"
            elif ext == ".csv":   source_type = "csv"
            elif ext in (".xlsx",".xls",".xlsm",".ods"): source_type = "excel"
            else: source_type = "web"
        else:
            source_type = "web"

    if source_type == "web":
        return scrape_site(source, use_scroll=use_scroll)
    elif source_type == "pdf":
        return parse_pdf(source)
    elif source_type in ("excel", "csv"):
        return parse_excel(source)
    return []


# ─────────────────────────────────────────────────────────────
#  USO DIRECTO
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python scraper.py <url_o_archivo> [--csv|--excel] [output]")
        sys.exit(0)

    src = sys.argv[1]
    fmt = "excel"
    out = f"precios_{datetime.now().strftime('%Y%m%d_%H%M')}"

    if "--csv" in sys.argv:
        fmt = "csv"
    if "--excel" in sys.argv:
        fmt = "excel"

    items = scrape(src)
    if not items:
        print("No se encontraron precios.")
        sys.exit(1)

    print(f"\n✅ {len(items)} productos encontrados\n")
    for i in items[:5]:
        print(f"  {i.get('nombre','')[:50]:<52} {i.get('precio_raw',''):>15}")
    if len(items) > 5:
        print(f"  ... y {len(items)-5} más")

    if fmt == "csv":
        export_csv(items, out + ".csv")
        print(f"\n📄 CSV guardado: {out}.csv")
    else:
        export_excel(items, out + ".xlsx")
        print(f"\n📊 Excel guardado: {out}.xlsx")
